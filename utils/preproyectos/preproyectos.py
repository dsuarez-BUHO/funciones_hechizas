"""
Lógica de dominio — Reporte de análisis de preproyectos.

Contiene las etapas de extracción, limpieza, agrupación y generación de reporte
específicas para el proceso de preproyectos de los clientes DQ, FAB y FDA.

Uso mínimo:
    from utils.preproyectos.preproyectos import run_pipeline_preproyectos

    URLS = { ... }
    HEADERS = { "api-key": "..." }
    df_estrategias, df_campanas, df_clientes = run_pipeline_preproyectos(URLS, HEADERS)
"""

import utils.pipeline_functions as Pipe
import utils.ETL_EDA_functions as EDA

import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows


# =============================================================================
# ETAPA 1 — EXTRACCIÓN
# =============================================================================

def fetch_all(urls, headers):
    """
    Hace GET a cada URL del dict usando Pipe.api_to_df y retorna los DataFrames crudos.
    Un DataFrame vacío indica fallo en esa llamada.

    Args:
        urls   : dict[str, str]  — {"cliente_año": "https://..."}
        headers: dict[str, str]  — headers HTTP (api-key, etc.)

    Returns:
        dict[str, pd.DataFrame]  — {"cliente_año": df} (vacío si falló)
    """
    resultados = {}
    for nombre, url in urls.items():
        df = Pipe.api_to_df(url, params=None, headers=headers, records_key="data")
        estado = "OK   " if not df.empty else "ERROR"
        print(f"  {estado} {nombre} | {len(df)} registros")
        resultados[nombre] = df
    return resultados


# =============================================================================
# ETAPA 2 — CONSTRUCCIÓN Y LIMPIEZA
# =============================================================================

def build_dataframe(resultados):
    """
    Concatena los DataFrames por fuente e infiere Cliente y Año desde la clave ("cliente_año").

    Args:
        resultados: dict[str, pd.DataFrame] — salida de fetch_all()

    Returns:
        pd.DataFrame — datos crudos con columnas Cliente y Año añadidas
    """
    dfs = []
    for nombre, df_temp in resultados.items():
        if df_temp.empty:
            continue
        df_temp = df_temp.copy()
        partes = nombre.split("_")
        df_temp["Cliente"] = partes[0].upper()
        df_temp["Año"]     = int(partes[1])
        dfs.append(df_temp)
    return pd.concat(dfs, ignore_index=True) if dfs else pd.DataFrame()


_COLUMNAS_ELIMINAR = [
    "arte", "nombre_arte", "nombre_usuario_artes", "usuario_artes",
    "preproyecto", "nombre_preproyecto", "nombre_usuario_preproyectos", "usuario_preproyectos"
]


def clean_dataframe(df):
    """
    Limpia el DataFrame crudo:
        - Elimina columnas irrelevantes
        - Convierte Folio a Int64
        - Parsea todas las columnas "fecha*" con EDA.auto_parse_dates + EDA.parse_date
        - Crea fecha_aprobacion = max(fecha_aprobado_arte, fecha_aprobado_ODT)
        - Reordena columnas poniendo Cliente y Folio al frente

    Args:
        df: pd.DataFrame — salida de build_dataframe()

    Returns:
        pd.DataFrame — datos limpios listos para análisis
    """
    df = df.drop(columns=_COLUMNAS_ELIMINAR, errors="ignore")

    df["Folio"] = pd.to_numeric(df["Folio"], errors="coerce").astype("Int64")

    # Parseo de fechas vía EDA (detecta automáticamente columnas con "fecha" en el nombre)
    df = EDA.auto_parse_dates(df, EDA.parse_date)

    # Elimina timezone — Excel no soporta datetimes con tz
    for col in df.select_dtypes(include=["datetimetz"]).columns:
        df[col] = df[col].dt.tz_convert(None)

    df["fecha_aprobacion"] = df[
        [c for c in ["fecha_aprobado_arte", "fecha_aprobado_ODT"] if c in df.columns]
    ].max(axis=1)
    df = df.drop(columns=["fecha_aprobado_arte", "fecha_aprobado_ODT"], errors="ignore")

    cols = df.columns.tolist()
    for target, pos in [("Folio", 2), ("Cliente", 1)]:
        if target in cols:
            cols.insert(pos, cols.pop(cols.index(target)))
    df = df[cols]

    return df


# =============================================================================
# ETAPA 3 — AGRUPACIÓN POR CAMPAÑA
# =============================================================================

_COLS_INVERTIDO_CAMPANA = [
    "duracion_carga_arte", "duracion_carga_preproyectos", "duracion_aprobacion"
]
_COLS_ETAPAS_CAMPANA = [
    "duracion_etapa_uno", "duracion_etapa_dos", "duracion_etapa_tres"
]


def build_campanas(df):
    """
    Agrupa el DataFrame de estrategias por [Cliente, Año, Campaña] y calcula:
        - Fechas mínimas y máximas de cada etapa
        - Duraciones en horas laborales (métrica principal) y timedelta (infraestructura)
        - Totales y porcentajes de tiempo

    Args:
        df: pd.DataFrame — salida de clean_dataframe()

    Returns:
        pd.DataFrame — una fila por campaña con todas las métricas
    """
    df_camp = df.drop(columns=["Folio", "Estrategia"], errors="ignore")

    df_camp = (
        df_camp
        .groupby(["Cliente", "Año", "Campaña"], as_index=False, dropna=False)
        .agg(
            total_estrategias    =("Campaña",         "size"),
            min_fecha_arte       =("fecha_arte",       "min"),
            max_fecha_arte       =("fecha_arte",       "max"),
            min_fecha_preproyecto=("fecha_preproyecto","min"),
            max_fecha_preproyecto=("fecha_preproyecto","max"),
            min_fecha_aprobacion =("fecha_aprobacion", "min"),
            max_fecha_aprobacion =("fecha_aprobacion", "max"),
        )
    )

    # Timedelta — infraestructura conservada para análisis futuro
    df_camp["duracion_carga_arte_td"]         = df_camp["max_fecha_arte"]        - df_camp["min_fecha_arte"]
    df_camp["duracion_carga_preproyectos_td"] = df_camp["max_fecha_preproyecto"] - df_camp["min_fecha_preproyecto"]
    df_camp["duracion_aprobacion_td"]         = df_camp["max_fecha_aprobacion"]  - df_camp["min_fecha_aprobacion"]
    df_camp["duracion_etapa_uno_td"]          = df_camp["max_fecha_arte"]        - df_camp["min_fecha_arte"]
    df_camp["duracion_etapa_dos_td"]          = df_camp["max_fecha_preproyecto"] - df_camp["max_fecha_arte"]
    df_camp["duracion_etapa_tres_td"]         = df_camp["max_fecha_aprobacion"]  - df_camp["max_fecha_preproyecto"]

    # Horas laborales — métrica principal
    df_camp["duracion_carga_arte"]         = EDA.calcular_horas_laborales(df_camp, "min_fecha_arte",        "max_fecha_arte")
    df_camp["duracion_carga_preproyectos"] = EDA.calcular_horas_laborales(df_camp, "min_fecha_preproyecto", "max_fecha_preproyecto")
    df_camp["duracion_aprobacion"]         = EDA.calcular_horas_laborales(df_camp, "min_fecha_aprobacion",  "max_fecha_aprobacion")
    df_camp["duracion_etapa_uno"]          = EDA.calcular_horas_laborales(df_camp, "min_fecha_arte",        "max_fecha_arte")
    df_camp["duracion_etapa_dos"]          = EDA.calcular_horas_laborales(df_camp, "max_fecha_arte",        "max_fecha_preproyecto")
    df_camp["duracion_etapa_tres"]         = EDA.calcular_horas_laborales(df_camp, "max_fecha_preproyecto", "max_fecha_aprobacion")

    df_camp = EDA.calcular_totales_y_porcentajes_bh(df_camp, _COLS_INVERTIDO_CAMPANA, _COLS_ETAPAS_CAMPANA)
    return df_camp


# =============================================================================
# ETAPA 4 — AGRUPACIÓN POR CLIENTE
# =============================================================================

_COLS_INVERTIDO_CLIENTE = [
    "promedio_duracion_carga_arte", "promedio_duracion_carga_preproyectos", "promedio_duracion_aprobacion"
]
_COLS_ETAPAS_CLIENTE = [
    "promedio_duracion_etapa_uno", "promedio_duracion_etapa_dos", "promedio_duracion_etapa_tres"
]


def build_clientes(df_campanas):
    """
    Agrupa el DataFrame de campañas por Cliente y calcula:
        - Total de campañas y estrategias
        - Promedios de todas las duraciones
        - Totales y porcentajes de tiempo promedio

    Args:
        df_campanas: pd.DataFrame — salida de build_campanas()

    Returns:
        pd.DataFrame — una fila por cliente con todas las métricas promedio
    """
    df_cli = (
        df_campanas
        .groupby("Cliente", as_index=False)
        .agg(
            total_estrategias                   =("total_estrategias",         "sum"),
            total_campanas                      =("Cliente",                   "count"),
            promedio_duracion_carga_arte        =("duracion_carga_arte",       "mean"),
            promedio_duracion_carga_preproyectos=("duracion_carga_preproyectos","mean"),
            promedio_duracion_aprobacion        =("duracion_aprobacion",       "mean"),
            promedio_duracion_etapa_uno         =("duracion_etapa_uno",        "mean"),
            promedio_duracion_etapa_dos         =("duracion_etapa_dos",        "mean"),
            promedio_duracion_etapa_tres        =("duracion_etapa_tres",       "mean"),
        )
    )
    df_cli = EDA.calcular_totales_y_porcentajes_bh(df_cli, _COLS_INVERTIDO_CLIENTE, _COLS_ETAPAS_CLIENTE)
    return df_cli


# =============================================================================
# ETAPA 5 — GENERACIÓN DEL REPORTE EXCEL
# =============================================================================

# ── Configuración de exportación ─────────────────────────────────────────────

# Columnas timedelta (_td) — infraestructura conservada para análisis futuro
_COLS_TD_CAMPANAS = [
    "duracion_carga_arte_td", "duracion_carga_preproyectos_td", "duracion_aprobacion_td",
    "duracion_etapa_uno_td", "duracion_etapa_dos_td", "duracion_etapa_tres_td",
]
_COLS_TD_CLIENTES = []  # los clientes ya no tienen columnas _td directas

# Columnas de horas laborales (float) — métrica principal
_COLS_BH_CAMPANAS = [
    "duracion_carga_arte", "duracion_carga_preproyectos", "duracion_aprobacion",
    "duracion_etapa_uno", "duracion_etapa_dos", "duracion_etapa_tres",
    "total_tiempo_invertido", "total_tiempo_campaña"
]
_COLS_BH_CLIENTES = [
    "promedio_duracion_carga_arte", "promedio_duracion_carga_preproyectos", "promedio_duracion_aprobacion",
    "promedio_duracion_etapa_uno", "promedio_duracion_etapa_dos", "promedio_duracion_etapa_tres",
    "total_tiempo_invertido", "total_tiempo_campaña"
]

_RENAME_CAMPANAS = {
    "Campaña": "Campaña", "Cliente": "Cliente", "total_estrategias": "Total de estrategias",
    "min_fecha_arte": "Primer carga de arte", "max_fecha_arte": "Última carga de arte",
    "min_fecha_preproyecto": "Primer carga de preproyecto", "max_fecha_preproyecto": "Última carga de preproyecto",
    "min_fecha_aprobacion": "Primer aprobación", "max_fecha_aprobacion": "Última aprobación",
    "duracion_carga_arte": "Total de carga de artes",
    "duracion_carga_preproyectos": "Total de carga de preproyectos",
    "duracion_aprobacion": "Total de aprobación",
    "duracion_etapa_uno": "Total Etapa 1", "duracion_etapa_dos": "Total Etapa 2",
    "duracion_etapa_tres": "Total Etapa 3",
    "total_tiempo_invertido": "Tiempo total invertido", "total_tiempo_campaña": "Tiempo total de campaña",
    "porcentaje_duracion_carga_arte": "Porcentaje duracion carga arte",
    "porcentaje_duracion_carga_preproyectos": "Porcentaje duracion carga preproyectos",
    "porcentaje_duracion_aprobacion": "Porcentaje duracion aprobacion",
    "porcentaje_duracion_etapa_uno": "Porcentaje duracion etapa 1",
    "porcentaje_duracion_etapa_dos": "Porcentaje duracion etapa 2",
    "porcentaje_duracion_etapa_tres": "Porcentaje duracion etapa 3",
}

_RENAME_CLIENTES = {
    "Cliente": "Cliente", "total_estrategias": "Total de estrategias",
    "total_campanas": "Total de campañas",
    "promedio_duracion_carga_arte": "Promedio de duración de carga de arte",
    "promedio_duracion_carga_preproyectos": "Promedio de duración de carga de preproyectos",
    "promedio_duracion_aprobacion": "Promedio de duración de aprobación",
    "promedio_duracion_etapa_uno": "Promedio Etapa 1", "promedio_duracion_etapa_dos": "Promedio Etapa 2",
    "promedio_duracion_etapa_tres": "Promedio Etapa 3",
    "total_tiempo_invertido": "Promedio total invertido", "total_tiempo_campaña": "Promedio total de campaña",
    "porcentaje_promedio_duracion_carga_arte": "Porcentaje promedio duracion carga arte",
    "porcentaje_promedio_duracion_carga_preproyectos": "Porcentaje promedio duracion carga preproyectos",
    "porcentaje_promedio_duracion_aprobacion": "Porcentaje promedio duracion aprobacion",
    "porcentaje_promedio_duracion_etapa_uno": "Porcentaje promedio duracion etapa 1",
    "porcentaje_promedio_duracion_etapa_dos": "Porcentaje promedio duracion etapa 2",
    "porcentaje_promedio_duracion_etapa_tres": "Porcentaje promedio duracion etapa 3",
}

_ORDEN_CLIENTES = [
    "Cliente", "Total de campañas", "Total de estrategias",
    "Promedio de duración de carga de arte", "Porcentaje promedio duracion carga arte",
    "Promedio de duración de carga de preproyectos", "Porcentaje promedio duracion carga preproyectos",
    "Promedio de duración de aprobación", "Porcentaje promedio duracion aprobacion",
    "Promedio total invertido",
    "Promedio Etapa 1", "Porcentaje promedio duracion etapa 1",
    "Promedio Etapa 2", "Porcentaje promedio duracion etapa 2",
    "Promedio Etapa 3", "Porcentaje promedio duracion etapa 3",
    "Promedio total de campaña",
]

_ORDEN_CAMPANAS = [
    "Cliente", "Campaña", "Total de estrategias",
    "Primer carga de arte", "Última carga de arte",
    "Primer carga de preproyecto", "Última carga de preproyecto",
    "Primer aprobación", "Última aprobación",
    "Total de carga de artes", "Porcentaje duracion carga arte",
    "Total de carga de preproyectos", "Porcentaje duracion carga preproyectos",
    "Total de aprobación", "Porcentaje duracion aprobacion",
    "Tiempo total invertido",
    "Total Etapa 1", "Porcentaje duracion etapa 1",
    "Total Etapa 2", "Porcentaje duracion etapa 2",
    "Total Etapa 3", "Porcentaje duracion etapa 3",
    "Tiempo total de campaña",
]

_PCT_CAMPANAS_EXPORT = {
    "Porcentaje duracion carga arte", "Porcentaje duracion carga preproyectos",
    "Porcentaje duracion aprobacion", "Porcentaje duracion etapa 1",
    "Porcentaje duracion etapa 2", "Porcentaje duracion etapa 3",
}

_PCT_CLIENTES_EXPORT = {
    "Porcentaje promedio duracion carga arte", "Porcentaje promedio duracion carga preproyectos",
    "Porcentaje promedio duracion aprobacion", "Porcentaje promedio duracion etapa 1",
    "Porcentaje promedio duracion etapa 2", "Porcentaje promedio duracion etapa 3",
}

_PASTEL = {"base": "D9EAD3", "fechas": "FCE5CD", "invertido": "D9EAF7", "etapas": "EADCF8"}

_HEADER_COLORS_CLIENTES = {
    "Cliente": _PASTEL["base"], "Total de campañas": _PASTEL["base"], "Total de estrategias": _PASTEL["base"],
    "Promedio de duración de carga de arte": _PASTEL["invertido"],
    "Porcentaje promedio duracion carga arte": _PASTEL["invertido"],
    "Promedio de duración de carga de preproyectos": _PASTEL["invertido"],
    "Porcentaje promedio duracion carga preproyectos": _PASTEL["invertido"],
    "Promedio de duración de aprobación": _PASTEL["invertido"],
    "Porcentaje promedio duracion aprobacion": _PASTEL["invertido"],
    "Promedio total invertido": _PASTEL["invertido"],
    "Promedio Etapa 1": _PASTEL["etapas"], "Porcentaje promedio duracion etapa 1": _PASTEL["etapas"],
    "Promedio Etapa 2": _PASTEL["etapas"], "Porcentaje promedio duracion etapa 2": _PASTEL["etapas"],
    "Promedio Etapa 3": _PASTEL["etapas"], "Porcentaje promedio duracion etapa 3": _PASTEL["etapas"],
    "Promedio total de campaña": _PASTEL["etapas"],
}

_HEADER_COLORS_CAMPANAS = {
    "Cliente": _PASTEL["base"], "Campaña": _PASTEL["base"], "Total de estrategias": _PASTEL["base"],
    "Primer carga de arte": _PASTEL["fechas"], "Última carga de arte": _PASTEL["fechas"],
    "Primer carga de preproyecto": _PASTEL["fechas"], "Última carga de preproyecto": _PASTEL["fechas"],
    "Primer aprobación": _PASTEL["fechas"], "Última aprobación": _PASTEL["fechas"],
    "Total de carga de artes": _PASTEL["invertido"],
    "Porcentaje duracion carga arte": _PASTEL["invertido"],
    "Total de carga de preproyectos": _PASTEL["invertido"],
    "Porcentaje duracion carga preproyectos": _PASTEL["invertido"],
    "Total de aprobación": _PASTEL["invertido"],
    "Porcentaje duracion aprobacion": _PASTEL["invertido"],
    "Tiempo total invertido": _PASTEL["invertido"],
    "Total Etapa 1": _PASTEL["etapas"], "Porcentaje duracion etapa 1": _PASTEL["etapas"],
    "Total Etapa 2": _PASTEL["etapas"], "Porcentaje duracion etapa 2": _PASTEL["etapas"],
    "Total Etapa 3": _PASTEL["etapas"], "Porcentaje duracion etapa 3": _PASTEL["etapas"],
    "Tiempo total de campaña": _PASTEL["etapas"],
}

_NOTAS = [
    "Los promedios del análisis global por cliente (tabla superior) pueden estar sesgados, ya que existen varias campañas emergentes con una sola estrategia o con muy pocas estrategias. Esto provoca duraciones de 0 minutos y reduce de forma importante los tiempos promedio. Es importante considerar este contexto al interpretar los resultados.",
    "Para futuras versiones del reporte, se recomienda aplicar un filtro que permita evaluar únicamente las campañas tradicionales o aquellas que realmente sea pertinente analizar. Esto puede resolverse definiendo un umbral mínimo de estrategias o mediante una validación manual.",
    "Se recomienda complementar siempre el análisis con la vista por campaña (tabla inferior), dando prioridad a la revisión de campañas tradicionales.",
    "Existen registros con duraciones negativas en la tabla \u201cAnálisis por campaña del cliente\u201d (filas marcadas en rojo). Ocurre cuando la última estrategia fue aprobada por la cuenta \u201cBot\u201d y la API no devuelve el registro de aprobación correspondiente.",
]

_DEFINICIONES = [
    ("Total de carga de artes",        "Tiempo total transcurrido desde la carga del primer arte hasta la carga del último arte."),
    ("Total de carga de preproyectos", "Tiempo total transcurrido desde la carga del primer preproyecto hasta la carga del último preproyecto."),
    ("Total de aprobación",            "Tiempo total transcurrido desde la primera aprobación hasta la última. Solo se consideran aprobaciones del cliente."),
    ("Total Etapa 1",                  "Tiempo total transcurrido desde la carga del primer arte hasta la carga del último arte."),
    ("Total Etapa 2",                  "Tiempo total transcurrido desde la carga del último arte hasta la carga del último preproyecto."),
    ("Total Etapa 3",                  "Tiempo total transcurrido desde la carga del último preproyecto hasta la última aprobación."),
    ("Tiempo total invertido",         "Suma de los tiempos de cada etapa medidos de forma independiente. Puede haber solapamientos entre etapas."),
    ("Tiempo total de campaña",        "Suma de Etapa 1 + Etapa 2 + Etapa 3. Tiempo real desde primer arte hasta última aprobación."),
    ("Promedio total invertido",       "Suma de los promedios de tiempo invertido por cliente."),
    ("Promedio total de campaña",      "Suma de los promedios de duración de etapas por cliente."),
]

_COLS_DURACION_ORIG = [
    "duracion_carga_arte", "duracion_carga_preproyectos", "duracion_aprobacion",
    "duracion_etapa_uno", "duracion_etapa_dos", "duracion_etapa_tres"
]


# ── Helpers de estilo Excel ───────────────────────────────────────────────────

def _sanitize_sheet_name(name, used_names):
    invalid = ['\\', '/', '*', '[', ']', ':', '?']
    clean = str(name)
    for ch in invalid:
        clean = clean.replace(ch, " ")
    clean = " ".join(clean.split()).strip() or "Cliente"
    clean = clean[:31]
    base, i = clean, 1
    while clean in used_names:
        suffix = f"_{i}"
        clean = base[:31 - len(suffix)] + suffix
        i += 1
    used_names.add(clean)
    return clean


def _style_header(cell, fill_color="D9EAD3"):
    cell.font      = Font(bold=True, color="000000")
    cell.fill      = PatternFill("solid", fgColor=fill_color)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _style_title(ws, row, start_col, end_col, text):
    ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=end_col)
    cell = ws.cell(row=row, column=start_col, value=text)
    cell.font      = Font(bold=True, color="FFFFFF", size=12)
    cell.fill      = PatternFill("solid", fgColor="0F243E")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 22


def _ajustar_anchos(ws, min_width=12, max_width=35):
    for col_cells in ws.columns:
        col_idx = col_cells[0].column
        max_len = max(
            (len(str(cell.value)) for cell in col_cells if cell.value is not None),
            default=0
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = max(min(max_len + 2, max_width), min_width)


def _escribir_dataframe(ws, df, start_row, title, percentage_columns=None, header_colors=None):
    percentage_columns = percentage_columns or set()
    header_colors      = header_colors or {}
    ncols = len(df.columns)
    if ncols == 0:
        return start_row, start_row, start_row, 0

    _style_title(ws, start_row, 1, ncols, title)
    header_row     = start_row + 1
    data_start_row = start_row + 2

    for j, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=header_row, column=j, value=col_name)
        _style_header(cell, fill_color=header_colors.get(col_name, "D9EAD3"))

    columnas_df = list(df.columns)
    for i, row in enumerate(dataframe_to_rows(df, index=False, header=False), start=data_start_row):
        for j, value in enumerate(row, start=1):
            col_name = columnas_df[j - 1]
            cell = ws.cell(row=i, column=j, value=value)
            if isinstance(value, (pd.Timestamp, datetime)):
                cell.number_format = "yyyy-mm-dd hh:mm:ss"
            if col_name in percentage_columns and value is not None and pd.notna(value):
                cell.number_format = "0.00%"
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    end_row = data_start_row + len(df) - 1 if len(df) > 0 else data_start_row
    return header_row, data_start_row, end_row, ncols


def generate_report(df_campanas, df_clientes, output_path="reporte_analisis_clientes.xlsx"):
    """
    Genera el archivo Excel con una hoja por cliente:
        - Tabla superior: análisis global del cliente
        - Tabla inferior: análisis por campaña (ordenado por fecha desc)
        - Filas rojas: campañas con duraciones negativas
        - Filas azules: campañas con 5+ estrategias

    Args:
        df_campanas : pd.DataFrame — salida de build_campanas()
        df_clientes : pd.DataFrame — salida de build_clientes()
        output_path : str          — ruta del archivo Excel a generar

    Returns:
        str — output_path confirmado
    """
    cols_pct_campanas = [c for c in df_campanas.columns if c.startswith("porcentaje_")]
    cols_pct_clientes = [c for c in df_clientes.columns if c.startswith("porcentaje_")]

    df_camp_exp = Pipe.preparar_para_export(df_campanas, _COLS_TD_CAMPANAS, cols_pct_campanas, cols_bh=_COLS_BH_CAMPANAS)
    df_cli_exp  = Pipe.preparar_para_export(df_clientes, _COLS_TD_CLIENTES, cols_pct_clientes, cols_bh=_COLS_BH_CLIENTES)

    df_camp_exp = df_camp_exp.rename(columns=_RENAME_CAMPANAS)
    df_cli_exp  = df_cli_exp.rename(columns=_RENAME_CLIENTES)

    df_camp_exp = df_camp_exp[[c for c in _ORDEN_CAMPANAS if c in df_camp_exp.columns]]
    df_cli_exp  = df_cli_exp[[c for c in _ORDEN_CLIENTES  if c in df_cli_exp.columns]]

    wb        = Workbook()
    ws_notas  = wb.active
    ws_notas.title = "Notas"

    fill_red         = PatternFill("solid", fgColor="F4CCCC")
    fill_blue        = PatternFill("solid", fgColor="CFE2F3")
    fill_dark        = PatternFill("solid", fgColor="0F243E")
    fill_section     = PatternFill("solid", fgColor="1F4E78")
    thin_gray        = Side(style="thin", color="D9D9D9")

    # Hoja Notas
    ws_notas.merge_cells("A1:E1")
    ws_notas["A1"] = "Notas del reporte"
    ws_notas["A1"].font      = Font(bold=True, color="FFFFFF", size=13)
    ws_notas["A1"].fill      = fill_dark
    ws_notas["A1"].alignment = Alignment(horizontal="center", vertical="center")

    fila = 3
    for i, nota in enumerate(_NOTAS, start=1):
        ws_notas.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=5)
        cell = ws_notas.cell(row=fila, column=1, value=f"{i}. {nota}")
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        cell.font = Font(size=11)
        ws_notas.row_dimensions[fila].height = 55
        fila += 1

    fila += 3
    ws_notas.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=5)
    cell = ws_notas.cell(row=fila, column=1, value="Definición de variables")
    cell.font = Font(bold=True, color="FFFFFF", size=12)
    cell.fill = fill_section
    cell.alignment = Alignment(horizontal="center", vertical="center")
    fila += 1

    ws_notas.cell(row=fila, column=1, value="Variable")
    ws_notas.cell(row=fila, column=2, value="Descripción")
    _style_header(ws_notas.cell(row=fila, column=1))
    _style_header(ws_notas.cell(row=fila, column=2))
    fila += 1

    for variable, descripcion in _DEFINICIONES:
        ws_notas.cell(row=fila, column=1, value=variable).alignment   = Alignment(wrap_text=True, vertical="top")
        ws_notas.cell(row=fila, column=2, value=descripcion).alignment = Alignment(wrap_text=True, vertical="top")
        fila += 1

    for col, width in [("A", 35), ("B", 110), ("C", 15), ("D", 15), ("E", 15)]:
        ws_notas.column_dimensions[col].width = width

    # Hojas por cliente
    clientes         = df_clientes["Cliente"].dropna().astype(str).drop_duplicates().tolist()
    used_sheet_names = {"Notas"}

    for cliente in clientes:
        sheet_name = _sanitize_sheet_name(cliente, used_sheet_names)
        ws = wb.create_sheet(title=sheet_name)

        df_global = df_cli_exp[df_cli_exp["Cliente"] == cliente].copy()
        _, _, end1, ncols1 = _escribir_dataframe(
            ws=ws, df=df_global, start_row=1, title="Análisis global del cliente",
            percentage_columns=_PCT_CLIENTES_EXPORT, header_colors=_HEADER_COLORS_CLIENTES
        )

        df_camp_src  = df_campanas[df_campanas["Cliente"] == cliente].copy()
        df_camp_src  = df_camp_src.sort_values("min_fecha_arte", ascending=False)
        df_camp_show = df_camp_exp[df_camp_exp["Cliente"] == cliente].copy()
        if len(df_camp_show) > 0:
            df_camp_show = df_camp_show.loc[df_camp_src.index].reset_index(drop=True)

        _, data_start2, end2, ncols2 = _escribir_dataframe(
            ws=ws, df=df_camp_show, start_row=end1 + 5, title="Análisis por campaña del cliente",
            percentage_columns=_PCT_CAMPANAS_EXPORT, header_colors=_HEADER_COLORS_CAMPANAS
        )

        for idx, (_, row) in enumerate(df_camp_src.iterrows()):
            tiene_negativo = any(
                pd.notna(row.get(col)) and row[col] < 0
                for col in _COLS_DURACION_ORIG if col in row.index
            )
            tiene_muchas = pd.notna(row.get("total_estrategias")) and row.get("total_estrategias") >= 5
            fill = fill_red if tiene_negativo else (fill_blue if tiene_muchas else None)
            if fill:
                for c in range(1, ncols2 + 1):
                    ws.cell(row=data_start2 + idx, column=c).fill = fill

        for row in ws.iter_rows(min_row=1, max_row=end2, min_col=1, max_col=max(ncols1, ncols2)):
            for cell in row:
                if cell.value is not None:
                    cell.border = Border(bottom=thin_gray)

        _ajustar_anchos(ws, min_width=12, max_width=32)

    wb.save(output_path)
    return output_path


# =============================================================================
# CLASIFICACIÓN EXTERNA — tradicional / extratemporal / sin clasificar
# =============================================================================

def cargar_clasificacion(path_excel, hoja='CHECK TRADICIONALES-STRATECO'):
    """
    Carga la hoja de clasificación del Excel externo y normaliza
    Cliente + Campaña para el join.

    Args:
        path_excel : str — ruta al archivo Excel de clasificación
        hoja       : str — nombre de la hoja (default: 'CHECK TRADICIONALES-STRATECO')

    Returns:
        pd.DataFrame — columnas: _join_cliente, _join_campana, tradicional
    """
    df = pd.read_excel(path_excel, sheet_name=hoja)
    df['_join_cliente'] = df['Cliente'].str.strip().str.upper()
    df['_join_campana'] = df['Campaña'].str.strip().str.upper()
    return df[['_join_cliente', '_join_campana', 'tradicional']].copy()


def _asignar_clasificacion_trad_ext(row, umbral=5):
    """
    Helper privado. Asigna clasificacion_campana_trad_ext a partir del join con Excel.

    Lógica (en cascada):
        1. Encontrada en Excel         → valor del Excel (0 ó 1)
        2. No encontrada + estr. > umbral → 1  (tradicional por umbral — pendiente validar con SME)
        3. No encontrada + estr. <= umbral → 2  (emergente / sin clasificar)

    Valores:
        1 = tradicional
        0 = extratemporal (>5 estrategias fuera de ventana de campaña)
        2 = emergente / sin clasificar
    """
    val = row['_clas_raw']
    if pd.notna(val):
        return int(val)
    if pd.notna(row['total_estrategias']) and row['total_estrategias'] > umbral:
        return 1
    return 2


def enriquecer_con_clasificacion(df_campanas_bq, df_clas, umbral=5):
    """
    Añade la columna clasificacion_campana_trad_ext a df_campanas_bq.

    Proceso:
        1. Left join por (cliente, campana) normalizado a mayúsculas
        2. Aplica _asignar_clasificacion_trad_ext() fila a fila
        3. Elimina columnas auxiliares del join
        4. Castea a Int64 (nullable — por si quedan NaN inesperados)

    Args:
        df_campanas_bq : pd.DataFrame — salida de build_preproyectos_campanas()
                         debe tener columnas 'cliente' y 'campana' (ya renombradas)
        df_clas        : pd.DataFrame — salida de cargar_clasificacion()
        umbral         : int          — estrategias mínimas para clasificar como
                         tradicional cuando no está en el Excel (default: 5)

    Returns:
        pd.DataFrame — df_campanas_bq con columna clasificacion_campana_trad_ext añadida
    """
    df = df_campanas_bq.copy()

    df['_join_cliente'] = df['cliente'].str.strip().str.upper()
    df['_join_campana'] = df['campana'].str.strip().str.upper()

    df = df.merge(
        df_clas.rename(columns={'tradicional': '_clas_raw'}),
        on=['_join_cliente', '_join_campana'],
        how='left'
    )

    df['clasificacion_campana_trad_ext'] = df.apply(
        lambda row: _asignar_clasificacion_trad_ext(row, umbral=umbral), axis=1
    )
    df['clasificacion_campana_trad_ext'] = df['clasificacion_campana_trad_ext'].astype('Int64')

    df = df.drop(columns=['_join_cliente', '_join_campana', '_clas_raw'])
    return df.reset_index(drop=True)


# =============================================================================
# PIPELINE PRINCIPAL
# =============================================================================

def run_pipeline_preproyectos(urls, headers, output_path="reporte_analisis_clientes.xlsx"):
    """
    Orquesta el pipeline completo de principio a fin.

    Args:
        urls       : dict[str, str] — {"cliente_año": "https://..."}
        headers    : dict[str, str] — headers HTTP (api-key, etc.)
        output_path: str            — ruta del Excel a generar

    Returns:
        tuple(df_estrategias, df_campanas, df_clientes)

    Ejemplo:
        URLS = {
            "dq_2025":  "https://dq.api.nest.buhoms.com/campaigns/get_str_report/2025",
            "dq_2026":  "https://dq.api.nest.buhoms.com/campaigns/get_str_report/2026",
            "fab_2025": "https://fab.api.nest.buhoms.com/campaigns/get_str_report/2025",
            "fab_2026": "https://fab.api.nest.buhoms.com/campaigns/get_str_report/2026",
            "fda_2025": "https://api.nest.buhoms.com/campaigns/get_str_report/2025",
            "fda_2026": "https://api.nest.buhoms.com/campaigns/get_str_report/2026",
        }
        HEADERS = {"api-key": "AQuvxn1YLf"}
        df_est, df_camp, df_cli = run_pipeline_preproyectos(URLS, HEADERS)
    """
    print("── [1/4] Extracción ────────────────────────────────")
    respuestas = fetch_all(urls, headers)

    print("── [2/4] Construcción del DataFrame ────────────────")
    df = build_dataframe(respuestas)
    df = clean_dataframe(df)
    df_estrategias = df.copy()
    print(f"  {len(df_estrategias)} estrategias cargadas")

    print("── [3/4] Agrupación ────────────────────────────────")
    df_campanas = build_campanas(df_estrategias)
    print(f"  {len(df_campanas)} campañas")
    df_clientes = build_clientes(df_campanas)
    print(f"  {len(df_clientes)} clientes")

    print("── [4/4] Reporte Excel ─────────────────────────────")
    generate_report(df_campanas, df_clientes, output_path)
    print(f"  Archivo generado: {output_path}")

    print("── Listo ────────────────────────────────────────────")
    return df_estrategias, df_campanas, df_clientes
